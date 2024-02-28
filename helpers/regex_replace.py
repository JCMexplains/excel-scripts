import argparse
import re

from openpyxl import load_workbook


def regex_replace(ws, find, replace):
    pattern = re.compile(fr'{find}')
    for row in ws.iter_rows():
        for cell in row:
            if cell.value and isinstance(cell.value, str):
                if pattern.search(cell.value):
                    cell.value = pattern.sub(replace, cell.value)


def main():
    parser = argparse.ArgumentParser(description='Search and replace text in an Excel worksheet using regular expressions.')
    parser.add_argument('filename', type=str, help='The Excel file to process.')
    parser.add_argument('find', type=str, help='The text to find (regular expression).')
    parser.add_argument('replace', type=str, help='The text to replace the found text with.')
    args = parser.parse_args()

    # Load the workbook and select the first worksheet
    wb = load_workbook(filename=args.filename)
    ws = wb.active

    # Perform the search and replace operation
    regex_replace(ws, args.find, args.replace)

    # Save the workbook
    wb.save(filename=args.filename)


if __name__ == "__main__":
    main()
