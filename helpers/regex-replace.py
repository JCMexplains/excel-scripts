import argparse
import re

from openpyxl import load_workbook


def regex_replace(ws, find, replace):
    pattern = re.compile(fr'{find}')
    for row in ws.iter_rows():
        for cell in row:
            # print(type(cell.value))
            if cell.value and isinstance(cell.value, str):
                if pattern.search(cell.value):
                    print('match: ' + cell.value)
                    cell.value = pattern.sub(replace, cell.value)
                    print('new: ' + cell.value)
            # elif cell.value:
                # print(cell.value)
                # temp_value = str(cell.value)
                # temp_value = pattern.sub(replace, temp_value)
                # cell.value = datetime.datetime(temp_value)


def main():
    parser = argparse.ArgumentParser(description='Search and replace text in an Excel worksheet using regular expressions.')
    parser.add_argument('filename', type=str, help='The Excel file to process.')
    parser.add_argument('find', type=str, help='The text to find (regular expression).')
    parser.add_argument('replace', type=str, help='The text to replace the found text with.')
    args = parser.parse_args()

    # Load the workbook and select the first worksheet
    wb = load_workbook(filename=args.filename)
    ws = wb.active

    # Perform the search and replace operation
    regex_replace(ws, args.find, args.replace)

    # Save the workbook
    wb.save(filename=args.filename)


if __name__ == "__main__":
    main()

# usage:
# py regex-replace.py filename.xlsx "text_to_find" "replacement_text"
