from openpyxl import load_workbook


def column_names_to_indices(file_path, sheet_name, column_names):
    '''
    Convert column names to indices in an Excel sheet.

    Parameters:
    - file_path: Path to the Excel file.
    - sheet_name: Name of the worksheet to search for column names.
    - column_names: A list of column names for which indices are required.

    Returns:
    A dictionary with column names as keys and their 1-based indices as values.
    '''
    # Load the worksheet directly
    ws = load_workbook(file_path)[sheet_name]

    # Find column indices in the third row
    col_indices = {cell_value: idx for idx, cell_value in enumerate(
        ws[3], start=1) if cell_value.value in column_names}

    return col_indices


# # Example usage
# file_path = 'data.xlsx'
# sheet_name = 'Sheet1'
# column_names = ['ColumnA', 'ColumnB']  # Example column names
# indices = column_names_to_indices(file_path, sheet_name, column_names)
# print(indices)

# # To print sorted indices in reverse order
# indices_list = sorted(indices.values(), reverse=True)
# print(indices_list)
