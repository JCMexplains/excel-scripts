from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo


def resize_table(ws, table_name):
    
    # removes table formatting, since otherwise 
    # deleting columns gives a weird error from a damaged table
    del ws.tables[table_name]

    # Determine the used range
    min_row = ws.min_row
    max_row = ws.max_row
    min_col = ws.min_column
    max_col = ws.max_column

    top_left = convert_rowcol_to_alpha(min_row, min_col)
    bottom_right = convert_rowcol_to_alpha(max_row, max_col)

    table = Table(displayName=table_name, ref=f'{top_left}:{bottom_right}')

    # Add a default table style with striped rows and banded columns
    style = TableStyleInfo(
        name='TableStyleMedium9', showFirstColumn=False,
        showLastColumn=False, showRowStripes=True, showColumnStripes=True
        )
    table.tableStyleInfo = style

    # Add the table to the worksheet
    ws.add_table(table)


def convert_rowcol_to_alpha(row, col):

    col_letter = get_column_letter(col)

    # Combine with a row number to get a full cell reference
    alpha = f'{col_letter}{row}'

    return alpha

    # print(alpha)  # This will print "B1"
