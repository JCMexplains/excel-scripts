from openpyxl import load_workbook


def convert_text_to_numbers_with_openpyxl(file_path):
    # Load the workbook and select the active worksheet
    wb = load_workbook(filename=file_path)
    ws = wb.active  
    # Assuming you want to work with the first sheet

    # Iterate through all rows and columns, 
    # converting text to numbers where applicable
    for row in ws.iter_rows():
        for cell in row:
            if cell.data_type == 's':  
                # Check if the cell data type is string ('s')
                try:
                    # Attempt to convert the string to a float
                    cell.value = float(cell.value)
                    cell.data_type = 'n'  # Set cell type to numeric ('n')
                except ValueError:
                    # If conversion fails, leave the value as is
                    continue

    # Save the modified workbook
    # Consider saving to a new file to preserve the original
    modified_file_path = "modified_" + file_path
    wb.save(filename=modified_file_path)


# Example usage
convert_text_to_numbers_with_openpyxl("data.xlsx")
